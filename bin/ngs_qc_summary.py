#!/usr/bin/env python3
"""
NGS QC 统计汇总脚本

从流程产生的 results/qc/ 目录中自动提取每个样本的质控信息，
汇总保存为 Excel 文件（含多 Sheet 和条件格式）。

目录结构要求：
  {results_dir}/qc/
    ├── fastp/{sample}/{sample}.json
    ├── bam/{sample}.flagstat.txt
    ├── bam/{sample}.stats.txt        (可选)
    ├── markdup/{sample}.markdup_metrics.txt  (可选)
    └── mosdepth/{sample}.mosdepth.summary.txt
               {sample}.mosdepth.global.dist.txt

用法：
  python3 ngs_qc_summary.py -i /path/to/results -o qc_summary.xlsx
  python3 ngs_qc_summary.py -i /path/to/results -o qc_summary.xlsx --depth-thresholds 10 20 30 50
"""

import os
import re
import json
import argparse
import sys
from typing import Optional

try:
    import pandas as pd
except ImportError:
    sys.exit("[ERROR] 请先安装 pandas: pip install pandas")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] 请先安装 openpyxl: pip install openpyxl")


# ─────────────────────────── 解析函数 ────────────────────────────

def parse_fastp(json_path: str) -> dict:
    """从 fastp JSON 提取过滤质控指标。"""
    with open(json_path) as f:
        d = json.load(f)

    bf = d["summary"]["before_filtering"]
    af = d["summary"]["after_filtering"]
    fr = d["filtering_result"]
    dup = d.get("duplication", {})

    return {
        "raw_reads":        bf["total_reads"],
        "raw_bases_Gb":     round(bf["total_bases"] / 1e9, 3),
        "clean_reads":      af["total_reads"],
        "clean_bases_Gb":   round(af["total_bases"] / 1e9, 3),
        "q20_before_%":     round(bf["q20_rate"] * 100, 2),
        "q30_before_%":     round(bf["q30_rate"] * 100, 2),
        "q20_after_%":      round(af["q20_rate"] * 100, 2),
        "q30_after_%":      round(af["q30_rate"] * 100, 2),
        "gc_before_%":      round(bf["gc_content"] * 100, 2),
        "filtered_reads":   bf["total_reads"] - af["total_reads"],
        "low_quality_reads": fr.get("low_quality_reads", 0),
        "too_many_N_reads":  fr.get("too_many_N_reads", 0),
        "too_short_reads":   fr.get("too_short_reads", 0),
        "adapter_trimmed_reads": d.get("adapter_cutting", {}).get("adapter_trimmed_reads", 0),
        "dup_rate_%":       round(dup.get("rate", 0) * 100, 4),
    }


def parse_flagstat(flagstat_path: str) -> dict:
    """从 samtools flagstat 提取比对统计。"""
    result = {}
    with open(flagstat_path) as f:
        text = f.read()

    # primary reads
    m = re.search(r"(\d+) \+ \d+ primary\b", text)
    if m:
        result["primary_reads"] = int(m.group(1))

    # primary mapped
    m = re.search(r"(\d+) \+ \d+ primary mapped \((.+?)%", text)
    if m:
        result["mapped_reads"]  = int(m.group(1))
        result["mapping_rate_%"] = float(m.group(2))

    # properly paired
    m = re.search(r"(\d+) \+ \d+ properly paired \((.+?)%", text)
    if m:
        result["properly_paired_reads"]  = int(m.group(1))
        result["properly_paired_rate_%"] = float(m.group(2))

    # primary duplicates
    m = re.search(r"(\d+) \+ \d+ primary duplicates", text)
    if m:
        result["dup_reads_flagstat"] = int(m.group(1))

    # singletons
    m = re.search(r"(\d+) \+ \d+ singletons \((.+?)%", text)
    if m:
        result["singleton_reads"]  = int(m.group(1))
        result["singleton_rate_%"] = float(m.group(2))

    return result


def parse_bam_stats(stats_path: str) -> dict:
    """从 samtools stats SN 行提取平均插入片段长度、错误率、平均质量等。"""
    result = {}
    keys = {
        "insert size average:": "insert_size_avg",
        "insert size standard deviation:": "insert_size_sd",
        "average quality:":  "avg_base_quality",
        "error rate:":       "error_rate",
        "reads MQ0:":        "reads_MQ0",
    }
    with open(stats_path) as f:
        for line in f:
            if not line.startswith("SN"):
                continue
            parts = line.split("\t")
            if len(parts) < 3:
                continue
            label = parts[1].strip()
            value = parts[2].split("#")[0].strip()
            if label in keys:
                try:
                    result[keys[label]] = float(value)
                except ValueError:
                    result[keys[label]] = value
    return result


def parse_mosdepth_summary(summary_path: str) -> dict:
    """从 mosdepth summary 提取全基因组平均深度。"""
    with open(summary_path) as f:
        for line in f:
            parts = line.strip().split("\t")
            if parts[0] == "total_region":
                return {"avg_depth_X": float(parts[3])}
    return {}


def parse_mosdepth_dist(dist_path: str, thresholds: list) -> dict:
    """从 mosdepth global dist 提取各深度阈值的覆盖率。"""
    depth_ratio = {}
    with open(dist_path) as f:
        for line in f:
            parts = line.strip().split("\t")
            if len(parts) == 3 and parts[0] == "total":
                d = int(parts[1])
                ratio = float(parts[2])
                if d in thresholds:
                    depth_ratio[d] = ratio

    result = {}
    for t in thresholds:
        result[f"cov_{t}X_%"] = round(depth_ratio.get(t, 0.0) * 100, 2)
    return result


def parse_markdup(markdup_path: str) -> dict:
    """从 sambamba markdup 日志提取重复 reads 数。"""
    result = {}
    with open(markdup_path) as f:
        text = f.read()
    m = re.search(r"found (\d+) duplicates", text)
    if m:
        result["dup_reads_markdup"] = int(m.group(1))
    return result


# ──────────────────────── 样本发现与汇总 ──────────────────────────

def discover_samples(qc_dir: str) -> list:
    """从 fastp 目录发现所有样本名称。"""
    fastp_dir = os.path.join(qc_dir, "fastp")
    if not os.path.isdir(fastp_dir):
        sys.exit(f"[ERROR] 未找到 fastp 目录: {fastp_dir}")
    samples = sorted(
        d for d in os.listdir(fastp_dir)
        if os.path.isdir(os.path.join(fastp_dir, d))
    )
    if not samples:
        sys.exit(f"[ERROR] fastp 目录下没有发现任何样本: {fastp_dir}")
    return samples


def collect_sample_qc(sample: str, qc_dir: str, depth_thresholds: list) -> dict:
    """聚合单个样本的所有 QC 指标。"""
    row = {"sample_id": sample}
    warnings = []

    # fastp
    fp = os.path.join(qc_dir, "fastp", sample, f"{sample}.json")
    if os.path.isfile(fp):
        row.update(parse_fastp(fp))
    else:
        warnings.append(f"  [WARN] 未找到 fastp JSON: {fp}")

    # flagstat
    fp = os.path.join(qc_dir, "bam", f"{sample}.flagstat.txt")
    if os.path.isfile(fp):
        row.update(parse_flagstat(fp))
    else:
        warnings.append(f"  [WARN] 未找到 flagstat: {fp}")

    # samtools stats
    fp = os.path.join(qc_dir, "bam", f"{sample}.stats.txt")
    if os.path.isfile(fp):
        row.update(parse_bam_stats(fp))
    else:
        warnings.append(f"  [WARN] 未找到 bam stats: {fp}")

    # mosdepth summary
    fp = os.path.join(qc_dir, "mosdepth", f"{sample}.mosdepth.summary.txt")
    if os.path.isfile(fp):
        row.update(parse_mosdepth_summary(fp))
    else:
        warnings.append(f"  [WARN] 未找到 mosdepth summary: {fp}")

    # mosdepth dist
    fp = os.path.join(qc_dir, "mosdepth", f"{sample}.mosdepth.global.dist.txt")
    if os.path.isfile(fp):
        row.update(parse_mosdepth_dist(fp, depth_thresholds))
    else:
        warnings.append(f"  [WARN] 未找到 mosdepth dist: {fp}")

    # markdup
    fp = os.path.join(qc_dir, "markdup", f"{sample}.markdup_metrics.txt")
    if os.path.isfile(fp):
        row.update(parse_markdup(fp))

    for w in warnings:
        print(w)

    return row


# ─────────────────────── Excel 写出 ──────────────────────────────

# 列分组：用于 Sheet 拆分和标题着色
COLUMN_GROUPS = {
    "基本信息": ["sample_id"],
    "fastp 过滤": [
        "raw_reads", "raw_bases_Gb",
        "clean_reads", "clean_bases_Gb",
        "q20_before_%", "q30_before_%",
        "q20_after_%", "q30_after_%",
        "gc_before_%",
        "filtered_reads", "low_quality_reads",
        "too_many_N_reads", "too_short_reads",
        "adapter_trimmed_reads", "dup_rate_%",
    ],
    "比对统计": [
        "primary_reads", "mapped_reads", "mapping_rate_%",
        "properly_paired_reads", "properly_paired_rate_%",
        "singleton_reads", "singleton_rate_%",
        "dup_reads_flagstat", "dup_reads_markdup",
    ],
    "BAM 质量": [
        "insert_size_avg", "insert_size_sd",
        "avg_base_quality", "error_rate", "reads_MQ0",
    ],
    "测序深度": ["avg_depth_X"],
}

# 中文列名映射
CN_NAMES = {
    "sample_id":               "样本ID",
    "raw_reads":               "原始Reads数",
    "raw_bases_Gb":            "原始数据量(Gb)",
    "clean_reads":             "过滤后Reads数",
    "clean_bases_Gb":          "过滤后数据量(Gb)",
    "q20_before_%":            "Q20过滤前(%)",
    "q30_before_%":            "Q30过滤前(%)",
    "q20_after_%":             "Q20过滤后(%)",
    "q30_after_%":             "Q30过滤后(%)",
    "gc_before_%":             "GC含量(%)",
    "filtered_reads":          "过滤Reads数",
    "low_quality_reads":       "低质量Reads",
    "too_many_N_reads":        "N超标Reads",
    "too_short_reads":         "过短Reads",
    "adapter_trimmed_reads":   "接头修剪Reads",
    "dup_rate_%":              "重复率fastp(%)",
    "primary_reads":           "Primary Reads",
    "mapped_reads":            "比对Reads数",
    "mapping_rate_%":          "比对率(%)",
    "properly_paired_reads":   "正确配对Reads",
    "properly_paired_rate_%":  "正确配对率(%)",
    "singleton_reads":         "Singleton Reads",
    "singleton_rate_%":        "Singleton率(%)",
    "dup_reads_flagstat":      "重复Reads数(flagstat)",
    "dup_reads_markdup":       "重复Reads数(markdup)",
    "insert_size_avg":         "平均插入片段(bp)",
    "insert_size_sd":          "插入片段SD(bp)",
    "avg_base_quality":        "平均碱基质量",
    "error_rate":              "错误率",
    "reads_MQ0":               "MQ0 Reads数",
    "avg_depth_X":             "平均测序深度(X)",
}

# 各组标题填充色（淡色）
GROUP_COLORS = {
    "基本信息":  "D9E1F2",
    "fastp 过滤": "E2EFDA",
    "比对统计":  "FCE4D6",
    "BAM 质量":  "FFF2CC",
    "测序深度":  "EDEDED",
}


def _border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _write_sheet(ws, df: pd.DataFrame, depth_thresholds: list):
    """将 DataFrame 写入一个工作表，并添加样式。"""
    # 按列分组确定表头颜色
    col_to_group = {}
    for grp, cols in COLUMN_GROUPS.items():
        for c in cols:
            col_to_group[c] = grp
    # 深度覆盖率列动态归入"测序深度"
    for t in depth_thresholds:
        col_to_group[f"cov_{t}X_%"] = "测序深度"
        CN_NAMES[f"cov_{t}X_%"] = f"≥{t}X覆盖率(%)"

    # 写表头
    headers = list(df.columns)
    for ci, col in enumerate(headers, start=1):
        grp = col_to_group.get(col, "基本信息")
        color = GROUP_COLORS.get(grp, "FFFFFF")
        cn = CN_NAMES.get(col, col)
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = Font(bold=True, size=10)
        cell.fill = _header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()

    # 写数据
    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
            if ri % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", fgColor="F7F7F7")

    # 自动列宽
    for ci, col in enumerate(headers, start=1):
        cn = CN_NAMES.get(col, col)
        max_len = max(len(str(cn)), 10)
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 28)

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"

    # 条件格式：比对率（绿/黄/红）
    if "mapping_rate_%" in headers:
        col_idx = headers.index("mapping_rate_%") + 1
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["99.5"],
                       fill=PatternFill(fill_type="solid", fgColor="C6EFCE")),
        )
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="lessThan", formula=["98"],
                       fill=PatternFill(fill_type="solid", fgColor="FFC7CE")),
        )

    # 条件格式：平均深度（色阶）
    if "avg_depth_X" in headers:
        col_idx = headers.index("avg_depth_X") + 1
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="FFC7CE",
                mid_type="num",   mid_value=30, mid_color="FFEB9C",
                end_type="max",   end_color="C6EFCE",
            ),
        )

    # 条件格式：Q30 过滤后（色阶）
    if "q30_after_%" in headers:
        col_idx = headers.index("q30_after_%") + 1
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="FFC7CE",
                mid_type="num",   mid_value=85, mid_color="FFEB9C",
                end_type="max",   end_color="C6EFCE",
            ),
        )


def save_excel(all_rows: list, output_path: str, depth_thresholds: list):
    """构建 DataFrame 并写入多 Sheet Excel。"""
    # 动态添加深度覆盖列到分组
    for t in depth_thresholds:
        if f"cov_{t}X_%" not in COLUMN_GROUPS["测序深度"]:
            COLUMN_GROUPS["测序深度"].append(f"cov_{t}X_%")

    # 统一列顺序
    ordered_cols = []
    for cols in COLUMN_GROUPS.values():
        ordered_cols.extend(cols)

    df_full = pd.DataFrame(all_rows)
    # 补齐缺失列
    for col in ordered_cols:
        if col not in df_full.columns:
            df_full[col] = None
    # 过滤实际存在的列
    final_cols = [c for c in ordered_cols if c in df_full.columns]
    df_full = df_full[final_cols]

    wb = openpyxl.Workbook()

    # Sheet 1：汇总（所有列）
    ws_all = wb.active
    ws_all.title = "QC汇总"
    _write_sheet(ws_all, df_full, depth_thresholds)

    # Sheet 2：fastp 质控
    fastp_cols = ["sample_id"] + [c for c in COLUMN_GROUPS["fastp 过滤"] if c in df_full.columns]
    ws_fastp = wb.create_sheet("fastp质控")
    _write_sheet(ws_fastp, df_full[fastp_cols], depth_thresholds)

    # Sheet 3：比对 & 深度
    align_cols = (
        ["sample_id"]
        + [c for c in COLUMN_GROUPS["比对统计"] if c in df_full.columns]
        + [c for c in COLUMN_GROUPS["BAM 质量"] if c in df_full.columns]
        + [c for c in COLUMN_GROUPS["测序深度"] if c in df_full.columns]
    )
    ws_align = wb.create_sheet("比对与深度")
    _write_sheet(ws_align, df_full[align_cols], depth_thresholds)

    wb.save(output_path)
    print(f"\n[OK] 结果已保存到: {output_path}")
    print(f"     包含 {len(all_rows)} 个样本，{len(final_cols)} 个指标")
    print(f"     Sheet: QC汇总 | fastp质控 | 比对与深度")


# ──────────────────────────── 主程序 ─────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="NGS QC 统计汇总脚本：提取 fastp/flagstat/mosdepth 质控信息并输出 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "-i", "--input", required=True,
        help="results 目录路径（包含 qc/ 子目录）",
    )
    parser.add_argument(
        "-o", "--output", default="qc_summary.xlsx",
        help="输出 Excel 文件路径（默认: qc_summary.xlsx）",
    )
    parser.add_argument(
        "--depth-thresholds", nargs="+", type=int,
        default=[1, 5, 10, 20, 30, 50],
        metavar="N",
        help="mosdepth 覆盖率统计阈值（默认: 1 5 10 20 30 50）",
    )
    parser.add_argument(
        "--qc-subdir", default="qc",
        help="qc 子目录名称（默认: qc）",
    )
    args = parser.parse_args()

    results_dir = os.path.abspath(args.input)
    if not os.path.isdir(results_dir):
        sys.exit(f"[ERROR] 输入目录不存在: {results_dir}")

    qc_dir = os.path.join(results_dir, args.qc_subdir)
    if not os.path.isdir(qc_dir):
        sys.exit(f"[ERROR] QC 目录不存在: {qc_dir}")

    print(f"[INFO] 扫描目录: {qc_dir}")
    samples = discover_samples(qc_dir)
    print(f"[INFO] 发现 {len(samples)} 个样本: {', '.join(samples)}\n")

    all_rows = []
    for sample in samples:
        print(f"[INFO] 处理样本: {sample}")
        row = collect_sample_qc(sample, qc_dir, args.depth_thresholds)
        all_rows.append(row)

    output_path = os.path.abspath(args.output)
    save_excel(all_rows, output_path, args.depth_thresholds)


if __name__ == "__main__":
    main()
